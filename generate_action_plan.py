"""
Generate professional Excel report for beastmode action plan.
Japanese language, formatted for non-technical stakeholders.
"""
import json, os, sys, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

DIRECTORY = r"d:\DOMO\2404"

# ── Styles ──
HEADER_FONT = Font(name='Meiryo', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
TITLE_FONT = Font(name='Meiryo', size=14, bold=True, color='2F5496')
SUBTITLE_FONT = Font(name='Meiryo', size=11, bold=True, color='2F5496')
CELL_FONT = Font(name='Meiryo', size=10)
CELL_FONT_SM = Font(name='Meiryo', size=9, color='555555')
PAGE_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ALT_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='center')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def style_cell(cell, font=CELL_FONT, fill=None, border=THIN_BORDER, alignment=None):
    cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if alignment: cell.alignment = alignment
    return cell


def extract_endpoint(har, page_id, pattern):
    for entry in har['log']['entries']:
        if pattern in entry['request']['url']:
            text = entry.get('response', {}).get('content', {}).get('text', '')
            if text:
                try: return json.loads(text)
                except: pass
    return None


def analyze_page(har_path, page_id):
    with open(har_path, 'r', encoding='utf-8') as f:
        har = json.load(f)
    analyzer = extract_endpoint(har, page_id, f'/pages/{page_id}/analyzer')
    stacks = extract_endpoint(har, page_id, f'/stacks/{page_id}/cards')
    if not analyzer or not stacks: return None

    page_title = stacks.get('page', {}).get('title', stacks.get('title', ''))
    data_sources = analyzer.get('details', {}).get('dataSources', [])
    if len(data_sources) <= 1: return None

    ds_info = {}
    for ds in data_sources:
        ds_id = ds.get('id', '')
        physical, formula = set(), set()
        for col in (ds.get('columns') or []):
            name = col.get('name', '')
            if not name or name.startswith('calculation_') or re.match(r'^[a-f0-9-]{36}$', name): continue
            if col.get('isCalculation') or col.get('classType') == 'FORMULA':
                formula.add(name)
            else:
                physical.add(name)
        for fm in (ds.get('formulas') or []):
            n = fm.get('name', '')
            if n and not n.startswith('calculation_'): formula.add(n)
        ds_info[ds_id] = {'name': ds.get('name', ''), 'physical': physical, 'formula': formula, 'all': physical | formula}

    cards = stacks.get('cards', [])
    card_ds = {}
    for card in cards:
        if card.get('type') != 'kpi': continue
        ct = card.get('metadata', {}).get('chartType', '')
        if any(x in ct for x in ['selector', 'slicer']): continue
        ds_ids = set()
        ds_names = []
        for d in card.get('datasources', []):
            did = d.get('dataSourceId', '')
            if did:
                ds_ids.add(did)
                ds_names.append(d.get('dataSourceName', did))
        card_ds[card.get('id', '')] = {'title': card.get('title', ''), 'ds_ids': ds_ids, 'ds_names': ds_names}

    # CAN recreate only
    results = []
    seen = set()
    for ci in card_ds.values():
        target_phys = set()
        target_all = set()
        for did in ci['ds_ids']:
            if did in ds_info:
                target_phys |= ds_info[did]['physical']
                target_all |= ds_info[did]['all']
        for src_id, src in ds_info.items():
            if src_id in ci['ds_ids']: continue
            if not (src['physical'] <= target_phys): continue
            for col in sorted(src['formula'] - target_all):
                key = (tuple(ci['ds_names']), src['name'], col)
                if key not in seen:
                    seen.add(key)
                    results.append({
                        'target': ', '.join(ci['ds_names']),
                        'source': src['name'],
                        'beastmode': col,
                    })
    return {'page_id': page_id, 'title': page_title, 'ds_info': ds_info, 'results': results}


def main():
    har_files = []
    for f in os.listdir(DIRECTORY):
        if f.endswith('.com') or f.endswith('.har'):
            m = re.search(r'page_(\d+)', f)
            if m: har_files.append((os.path.join(DIRECTORY, f), m.group(1)))
    har_files.sort(key=lambda x: int(x[1]))

    pages = []
    for hp, pid in har_files:
        try:
            pd = analyze_page(hp, pid)
            if pd and pd['results']:
                pages.append(pd)
        except Exception as e:
            print(f"ERROR {pid}: {e}")

    # ── Build Excel ──
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════
    # Sheet 1: サマリー (Summary)
    # ════════════════════════════════════════════
    ws = wb.active
    ws.title = 'サマリー'
    ws.sheet_properties.tabColor = '2F5496'

    # Title
    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = 'Filter Views 対応計画 ― Beastmode追加一覧'
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:D2')
    ws['A2'].value = '各ページのFilter Viewsで反応しないカードに対し、追加作成が可能なBeastmode（計算フィールド）の一覧です。'
    ws['A2'].font = CELL_FONT_SM
    ws.row_dimensions[2].height = 22

    # Summary table
    row = 4
    headers = ['No.', 'ページ名', 'Page ID', '追加Beastmode数']
    for c_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=c_idx, value=h)
        style_cell(c, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)

    total = 0
    for i, pd in enumerate(pages, 1):
        row += 1
        ws.cell(row=row, column=1, value=i).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=pd['title']).font = CELL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=int(pd['page_id'])).font = CELL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        cnt = len(pd['results'])
        ws.cell(row=row, column=4, value=cnt).font = Font(name='Meiryo', size=11, bold=True)
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = GREEN_FILL
        total += cnt

    row += 1
    ws.cell(row=row, column=2, value='合計').font = Font(name='Meiryo', size=11, bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=4, value=total).font = Font(name='Meiryo', size=12, bold=True, color='2F5496')
    ws.cell(row=row, column=4).alignment = CENTER_ALIGN
    ws.cell(row=row, column=4).border = THIN_BORDER
    ws.cell(row=row, column=4).fill = GREEN_FILL

    # Notes
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = '【判定条件】'
    ws[f'A{row}'].font = Font(name='Meiryo', size=10, bold=True)
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = 'Beastmodeの元となるDatasetの物理カラムがすべて対象Datasetにも存在する場合、同じ計算式で再作成が可能と判定しています。'
    ws[f'A{row}'].font = CELL_FONT_SM

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

    # ════════════════════════════════════════════
    # Sheet 2: 詳細一覧 (Detail list)
    # ════════════════════════════════════════════
    ws2 = wb.create_sheet('Beastmode詳細一覧')
    ws2.sheet_properties.tabColor = '548235'

    headers = ['No.', 'ページ名', 'Page ID', '対象Dataset（追加先）', 'コピー元Dataset', '追加するBeastmode名']
    for c_idx, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=c_idx, value=h)
        style_cell(c, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = 'A2'

    row = 2
    num = 0
    for pd in pages:
        # Group by (target, source)
        groups = {}
        for r in pd['results']:
            key = (r['target'], r['source'])
            groups.setdefault(key, []).append(r)

        for (tgt, src), items in groups.items():
            for r in items:
                num += 1
                fill = ALT_FILL if (num % 2 == 0) else None
                ws2.cell(row=row, column=1, value=num)
                ws2.cell(row=row, column=2, value=pd['title'])
                ws2.cell(row=row, column=3, value=int(pd['page_id']))
                ws2.cell(row=row, column=4, value=tgt)
                ws2.cell(row=row, column=5, value=src)
                ws2.cell(row=row, column=6, value=r['beastmode'])
                for ci in range(1, 7):
                    c = ws2.cell(row=row, column=ci)
                    c.font = CELL_FONT
                    c.border = THIN_BORDER
                    if fill: c.fill = fill
                    if ci in (1, 3): c.alignment = CENTER_ALIGN
                row += 1

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 35
    ws2.column_dimensions['E'].width = 35
    ws2.column_dimensions['F'].width = 60

    # Auto-filter
    ws2.auto_filter.ref = f'A1:F{row-1}'

    # ════════════════════════════════════════════
    # Per-page sheets
    # ════════════════════════════════════════════
    for pd in pages:
        sheet_name = pd['title'][:28]  # max 31 chars
        ws_p = wb.create_sheet(sheet_name)
        ws_p.sheet_properties.tabColor = 'BF8F00'

        # Title
        ws_p.merge_cells('A1:C1')
        ws_p['A1'].value = f"{pd['title']}（Page {pd['page_id']}）"
        ws_p['A1'].font = TITLE_FONT
        ws_p.row_dimensions[1].height = 30

        # Dataset overview
        ws_p.merge_cells('A3:C3')
        ws_p['A3'].value = '■ このページのDataset構成'
        ws_p['A3'].font = SUBTITLE_FONT
        
        ds_headers = ['Dataset名', '物理カラム数', 'Beastmode数']
        for ci, h in enumerate(ds_headers, 1):
            c = ws_p.cell(row=4, column=ci, value=h)
            style_cell(c, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
        
        r = 5
        for ds_id, info in pd['ds_info'].items():
            ws_p.cell(row=r, column=1, value=info['name']).font = CELL_FONT
            ws_p.cell(row=r, column=1).border = THIN_BORDER
            ws_p.cell(row=r, column=2, value=len(info['physical'])).font = CELL_FONT
            ws_p.cell(row=r, column=2).alignment = CENTER_ALIGN
            ws_p.cell(row=r, column=2).border = THIN_BORDER
            ws_p.cell(row=r, column=3, value=len(info['formula'])).font = CELL_FONT
            ws_p.cell(row=r, column=3).alignment = CENTER_ALIGN
            ws_p.cell(row=r, column=3).border = THIN_BORDER
            r += 1

        # Action items
        r += 1
        ws_p.merge_cells(f'A{r}:C{r}')
        ws_p[f'A{r}'].value = '■ 追加が必要なBeastmode一覧'
        ws_p[f'A{r}'].font = SUBTITLE_FONT
        r += 1

        groups = {}
        for res in pd['results']:
            key = (res['target'], res['source'])
            groups.setdefault(key, []).append(res)

        for (tgt, src), items in groups.items():
            ws_p.merge_cells(f'A{r}:C{r}')
            ws_p[f'A{r}'].value = f'▸ 追加先: {tgt} ← コピー元: {src}'
            ws_p[f'A{r}'].font = Font(name='Meiryo', size=10, bold=True, color='548235')
            ws_p[f'A{r}'].fill = GREEN_FILL
            ws_p[f'A{r}'].border = THIN_BORDER
            r += 1

            act_headers = ['No.', '追加するBeastmode名', 'ステータス']
            for ci, h in enumerate(act_headers, 1):
                c = ws_p.cell(row=r, column=ci, value=h)
                style_cell(c, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
            r += 1

            for idx, item in enumerate(items, 1):
                fill = ALT_FILL if idx % 2 == 0 else None
                ws_p.cell(row=r, column=1, value=idx).font = CELL_FONT
                ws_p.cell(row=r, column=1).alignment = CENTER_ALIGN
                ws_p.cell(row=r, column=1).border = THIN_BORDER
                if fill: ws_p.cell(row=r, column=1).fill = fill
                
                ws_p.cell(row=r, column=2, value=item['beastmode']).font = CELL_FONT
                ws_p.cell(row=r, column=2).border = THIN_BORDER
                if fill: ws_p.cell(row=r, column=2).fill = fill
                
                ws_p.cell(row=r, column=3, value='未対応').font = Font(name='Meiryo', size=9, color='C00000')
                ws_p.cell(row=r, column=3).alignment = CENTER_ALIGN
                ws_p.cell(row=r, column=3).border = THIN_BORDER
                if fill: ws_p.cell(row=r, column=3).fill = fill
                r += 1
            r += 1

        ws_p.column_dimensions['A'].width = 8
        ws_p.column_dimensions['B'].width = 60
        ws_p.column_dimensions['C'].width = 15

    # Save
    out_path = os.path.join(DIRECTORY, 'beastmode_action_plan.xlsx')
    wb.save(out_path)
    print(f"Excel saved: {out_path}")
    print(f"Total: {total} beastmodes across {len(pages)} pages")


if __name__ == '__main__':
    main()
