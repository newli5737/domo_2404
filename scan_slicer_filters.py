"""
Slicer Filter Cross-Dataset Audit
===================================
For each page:
1. Find all slicer cards (dropdown, checkbox, slicer widgets)
2. Extract the column each slicer filters on
3. For each non-slicer card, check if its dataset has that column
4. Output: CSV + XLSX report

Uses HAR files containing stacks + analyzer endpoints.
"""
import json, os, sys, re, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

DIRECTORY = r"d:\DOMO\2404"

# ── Styles ──
HEADER_FONT = Font(name='Meiryo', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
TITLE_FONT = Font(name='Meiryo', size=14, bold=True, color='2F5496')
SUBTITLE_FONT = Font(name='Meiryo', size=11, bold=True, color='2F5496')
CELL_FONT = Font(name='Meiryo', size=10)
CELL_FONT_SM = Font(name='Meiryo', size=9, color='555555')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
RED_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
ALT_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
SLICER_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
CENTER = Alignment(horizontal='center', vertical='center')


def load_har(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def extract_endpoint(har, pattern):
    for entry in har['log']['entries']:
        if pattern in entry['request']['url']:
            text = entry.get('response', {}).get('content', {}).get('text', '')
            if text:
                try: return json.loads(text)
                except: pass
    return None


def analyze_page(har_path, page_id):
    har = load_har(har_path)
    stacks = extract_endpoint(har, f'/stacks/{page_id}/cards')
    analyzer = extract_endpoint(har, f'/pages/{page_id}/analyzer')
    if not stacks: return None

    page_title = stacks.get('page', {}).get('title', stacks.get('title', f'Page {page_id}'))
    cards = stacks.get('cards', [])

    # ── Build dataset column map + ID→name map from analyzer ──
    ds_columns = {}   # ds_id -> set of column names (physical + formula)
    ds_names = {}     # ds_id -> ds_name
    col_id_to_name = {}  # calculation_xxx or column_id -> actual readable name
    if analyzer:
        for ds in analyzer.get('details', {}).get('dataSources', []):
            ds_id = ds.get('id', '')
            ds_names[ds_id] = ds.get('name', '')
            cols = set()
            for col in (ds.get('columns') or []):
                name = col.get('name', '')
                col_id = col.get('id', '')
                if name and not name.startswith('calculation_') and not re.match(r'^[a-f0-9-]{36}$', name):
                    cols.add(name)
                # Map col_id -> name (physical columns: id is usually same as name)
                if col_id and name:
                    col_id_to_name[col_id] = name
            for fm in (ds.get('formulas') or []):
                n = fm.get('name', '')
                fm_id = fm.get('id', '')
                if n and not n.startswith('calculation_'):
                    cols.add(n)
                # Map formula id (calculation_xxx) -> readable name
                if fm_id and n:
                    col_id_to_name[fm_id] = n
            ds_columns[ds_id] = cols

    # ── Identify slicer cards ──
    SLICER_TYPES = {'badge_slicer', 'badge_dropdown_selector', 'badge_checkbox_selector',
                    'badge_radio_button_selector', 'badge_date_selector', 'badge_range_selector'}
    slicers = []
    for card in cards:
        if card.get('type') != 'kpi': continue
        chart_type = card.get('metadata', {}).get('chartType', '')
        if chart_type not in SLICER_TYPES and 'selector' not in chart_type and 'slicer' not in chart_type:
            continue

        # Extract filter column ID
        filter_col_id = None
        for sub_item in card.get('subscriptions', []):
            inner = sub_item.get('subscription', {})
            for col in inner.get('columns', []):
                if col.get('mapping') == 'ITEM':
                    filter_col_id = col.get('column') or col.get('formulaId', '')
                    break
            if filter_col_id: break

        # For date selectors
        if not filter_col_id and 'date' in chart_type:
            for sub_item in card.get('subscriptions', []):
                inner = sub_item.get('subscription', {})
                dg = inner.get('dateGrain', {})
                if dg.get('column'):
                    filter_col_id = dg['column']
                    break

        if not filter_col_id: continue

        # Resolve actual column name:
        # Priority: analyzer col_id_to_name > raw col_id (if it's a readable name) > card title
        if filter_col_id in col_id_to_name:
            # calculation_xxx -> actual name like ユニット
            filter_col_name = col_id_to_name[filter_col_id]
        elif not filter_col_id.startswith('calculation_') and not re.match(r'^[a-f0-9-]{36}$', filter_col_id):
            # Already a readable name like 担当施工店
            filter_col_name = filter_col_id
        else:
            # Fallback to card title (should rarely happen)
            filter_col_name = card.get('title', filter_col_id)

        # Get slicer's own dataset
        slicer_ds_ids = set()
        slicer_ds_name = ''
        for ds in card.get('datasources', []):
            did = ds.get('dataSourceId', '')
            if did:
                slicer_ds_ids.add(did)
                slicer_ds_name = ds.get('dataSourceName', did)

        slicers.append({
            'card_id': card.get('id', ''),
            'title': card.get('title', ''),
            'chart_type': chart_type,
            'filter_col_id': filter_col_id,
            'filter_col_name': filter_col_name,
            'ds_ids': slicer_ds_ids,
            'ds_name': slicer_ds_name,
        })

    if not slicers:
        return None

    # ── Check each non-slicer card against each slicer ──
    results = []
    for card in cards:
        if card.get('type') != 'kpi': continue
        chart_type = card.get('metadata', {}).get('chartType', '')
        # Skip slicer cards themselves
        if chart_type in SLICER_TYPES or 'selector' in chart_type or 'slicer' in chart_type:
            continue

        card_id = card.get('id', '')
        card_title = card.get('title', '')
        card_ds_ids = set()
        card_ds_names = []
        for ds in card.get('datasources', []):
            did = ds.get('dataSourceId', '')
            if did:
                card_ds_ids.add(did)
                card_ds_names.append(ds.get('dataSourceName', did))

        # Get all columns available to this card
        card_all_cols = set()
        for ds_id in card_ds_ids:
            if ds_id in ds_columns:
                card_all_cols |= ds_columns[ds_id]

        for slicer in slicers:
            # Check: does card's dataset have the slicer's filter column?
            col_name = slicer['filter_col_name']
            col_id = slicer['filter_col_id']

            # Check by name match
            has_col = col_name in card_all_cols or col_id in card_all_cols

            # Also check if card shares dataset with slicer (same DS = definitely responds)
            shares_ds = bool(card_ds_ids & slicer['ds_ids'])

            responds = has_col or shares_ds

            results.append({
                'page_id': page_id,
                'page_title': page_title,
                'slicer_title': slicer['title'],
                'slicer_column': col_name,
                'slicer_dataset': slicer['ds_name'],
                'card_id': card_id,
                'card_title': card_title,
                'card_dataset': ', '.join(card_ds_names),
                'responds': '✅' if responds else '❌',
                'reason': ('同じDataset' if shares_ds else 'カラム有り') if responds else 'カラム無し',
            })

    return {
        'page_id': page_id, 'page_title': page_title,
        'slicers': slicers, 'results': results,
        'ds_names': ds_names,
    }


def main():
    har_files = []
    for f in os.listdir(DIRECTORY):
        if f.endswith('.com') or f.endswith('.har'):
            m = re.search(r'page_(\d+)', f)
            if m: har_files.append((os.path.join(DIRECTORY, f), m.group(1)))
    har_files.sort(key=lambda x: int(x[1]))

    all_pages = []
    all_rows = []
    for hp, pid in har_files:
        try:
            pd = analyze_page(hp, pid)
            if pd:
                all_pages.append(pd)
                all_rows.extend(pd['results'])
                no_respond = sum(1 for r in pd['results'] if r['responds'] == '❌')
                print(f"Page {pid} ({pd['page_title']}): {len(pd['slicers'])} slicers, "
                      f"{len(pd['results'])} checks, {no_respond} ❌")
            else:
                print(f"Page {pid}: No slicers found")
        except Exception as e:
            print(f"ERROR {pid}: {e}")
            import traceback; traceback.print_exc()

    # ── CSV ──
    csv_path = os.path.join(DIRECTORY, "slicer_filter_audit.csv")
    csv_keys = ['page_id', 'page_title', 'slicer_title', 'slicer_column', 'slicer_dataset',
                'card_id', 'card_title', 'card_dataset', 'responds', 'reason']
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=csv_keys)
        w.writeheader()
        for r in all_rows: w.writerow(r)
    print(f"\nCSV: {csv_path} ({len(all_rows)} rows)")

    total_no = sum(1 for r in all_rows if r['responds'] == '❌')
    total_yes = sum(1 for r in all_rows if r['responds'] == '✅')
    print(f"✅ Responds: {total_yes}")
    print(f"❌ Not responds: {total_no}")

    # ── Excel ──
    wb = openpyxl.Workbook()

    # Sheet 1: サマリー
    ws = wb.active
    ws.title = 'サマリー'
    ws.sheet_properties.tabColor = '2F5496'

    ws.merge_cells('A1:E1')
    ws['A1'].value = 'Slicer Filter 診断レポート ― カード応答確認'
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:E2')
    ws['A2'].value = '各ページのSlicerカード（フィルター）に対し、各カードが正しく応答するかを確認した結果です。'
    ws['A2'].font = CELL_FONT_SM

    # Summary table
    row = 4
    headers = ['No.', 'ページ名', 'Page ID', 'Slicer数', '未応答カード数']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER

    total_problems = 0
    for i, pd in enumerate(all_pages, 1):
        row += 1
        n_slicers = len(pd['slicers'])
        n_no = sum(1 for r in pd['results'] if r['responds'] == '❌')
        total_problems += n_no

        ws.cell(row=row, column=1, value=i).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = CENTER; ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=pd['page_title']).font = CELL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=int(pd['page_id'])).font = CELL_FONT
        ws.cell(row=row, column=3).alignment = CENTER; ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=n_slicers).font = CELL_FONT
        ws.cell(row=row, column=4).alignment = CENTER; ws.cell(row=row, column=4).border = THIN_BORDER
        c5 = ws.cell(row=row, column=5, value=n_no)
        c5.font = Font(name='Meiryo', size=11, bold=True, color='C00000' if n_no > 0 else '107C10')
        c5.alignment = CENTER; c5.border = THIN_BORDER
        c5.fill = RED_FILL if n_no > 0 else GREEN_FILL

    row += 1
    ws.cell(row=row, column=2, value='合計').font = Font(name='Meiryo', size=11, bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=4).border = THIN_BORDER
    c5 = ws.cell(row=row, column=5, value=total_problems)
    c5.font = Font(name='Meiryo', size=12, bold=True, color='C00000' if total_problems > 0 else '107C10')
    c5.alignment = CENTER; c5.border = THIN_BORDER

    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].value = '【凡例】 ✅ = Slicerのフィルターに正しく応答する ／ ❌ = Slicerのカラムが無いため応答しない'
    ws[f'A{row}'].font = CELL_FONT_SM

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    # Sheet 2: 未応答一覧 (only ❌)
    ws2 = wb.create_sheet('未応答カード一覧')
    ws2.sheet_properties.tabColor = 'C00000'

    no_rows = [r for r in all_rows if r['responds'] == '❌']
    headers = ['No.', 'ページ名', 'Slicer名', 'フィルターカラム', 'Slicerの Dataset',
               'カード名', 'カードのDataset', '状態']
    for ci, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = 'A2'

    for i, r in enumerate(no_rows, 1):
        row = i + 1
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [i, r['page_title'], r['slicer_title'], r['slicer_column'], r['slicer_dataset'],
                r['card_title'], r['card_dataset'], '❌ カラム無し']
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=ci, value=v)
            c.font = CELL_FONT; c.border = THIN_BORDER
            if fill: c.fill = fill
            if ci in (1, 8): c.alignment = CENTER

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 30
    ws2.column_dimensions['F'].width = 40
    ws2.column_dimensions['G'].width = 30
    ws2.column_dimensions['H'].width = 14
    if no_rows:
        ws2.auto_filter.ref = f'A1:H{len(no_rows)+1}'

    # Sheet 3: 全カード診断結果
    ws3 = wb.create_sheet('全カード診断結果')
    ws3.sheet_properties.tabColor = '548235'

    headers = ['No.', 'ページ名', 'Slicer名', 'フィルターカラム',
               'カード名', 'カードのDataset', '応答', '理由']
    for ci, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    ws3.row_dimensions[1].height = 30
    ws3.freeze_panes = 'A2'

    for i, r in enumerate(all_rows, 1):
        row = i + 1
        is_ok = r['responds'] == '✅'
        vals = [i, r['page_title'], r['slicer_title'], r['slicer_column'],
                r['card_title'], r['card_dataset'], r['responds'], r['reason']]
        for ci, v in enumerate(vals, 1):
            c = ws3.cell(row=row, column=ci, value=v)
            c.font = CELL_FONT; c.border = THIN_BORDER
            if ci in (1, 7): c.alignment = CENTER
            if ci == 7:
                c.fill = GREEN_FILL if is_ok else RED_FILL
                c.font = Font(name='Meiryo', size=10, bold=True,
                              color='107C10' if is_ok else 'C00000')

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['D'].width = 25
    ws3.column_dimensions['E'].width = 40
    ws3.column_dimensions['F'].width = 30
    ws3.column_dimensions['G'].width = 8
    ws3.column_dimensions['H'].width = 14
    if all_rows:
        ws3.auto_filter.ref = f'A1:H{len(all_rows)+1}'

    # Per-page sheets
    for pd in all_pages:
        n_no = sum(1 for r in pd['results'] if r['responds'] == '❌')
        if n_no == 0: continue  # Skip pages with no problems

        sheet_name = pd['page_title'][:28]
        ws_p = wb.create_sheet(sheet_name)
        ws_p.sheet_properties.tabColor = 'BF8F00'

        ws_p.merge_cells('A1:D1')
        ws_p['A1'].value = f"{pd['page_title']}（Page {pd['page_id']}）"
        ws_p['A1'].font = TITLE_FONT
        ws_p.row_dimensions[1].height = 30

        # Slicers overview
        ws_p.merge_cells('A3:D3')
        ws_p['A3'].value = '■ このページのSlicerフィルター'
        ws_p['A3'].font = SUBTITLE_FONT

        sh = ['Slicer名', 'フィルターカラム', 'Dataset', 'タイプ']
        for ci, h in enumerate(sh, 1):
            c = ws_p.cell(row=4, column=ci, value=h)
            c.font = HEADER_FONT; c.fill = SLICER_FILL
            c.font = Font(name='Meiryo', size=10, bold=True, color='BF8F00')
            c.border = THIN_BORDER
        r = 5
        for s in pd['slicers']:
            ws_p.cell(row=r, column=1, value=s['title']).font = CELL_FONT
            ws_p.cell(row=r, column=1).border = THIN_BORDER
            ws_p.cell(row=r, column=2, value=s['filter_col_name']).font = CELL_FONT
            ws_p.cell(row=r, column=2).border = THIN_BORDER
            ws_p.cell(row=r, column=3, value=s['ds_name']).font = CELL_FONT
            ws_p.cell(row=r, column=3).border = THIN_BORDER
            ws_p.cell(row=r, column=4, value=s['chart_type'].replace('badge_', '')).font = CELL_FONT_SM
            ws_p.cell(row=r, column=4).border = THIN_BORDER
            r += 1

        # Problem cards
        r += 1
        ws_p.merge_cells(f'A{r}:D{r}')
        ws_p[f'A{r}'].value = f'■ 未応答カード一覧（{n_no}件）'
        ws_p[f'A{r}'].font = SUBTITLE_FONT
        r += 1

        ph = ['カード名', 'カードのDataset', '未応答Slicer', 'フィルターカラム']
        for ci, h in enumerate(ph, 1):
            c = ws_p.cell(row=r, column=ci, value=h)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        r += 1

        # Group by card
        card_problems = {}
        for res in pd['results']:
            if res['responds'] == '❌':
                key = (res['card_title'], res['card_dataset'])
                card_problems.setdefault(key, []).append(res)

        idx = 0
        for (ct, cd), issues in card_problems.items():
            for iss in issues:
                idx += 1
                fill = ALT_FILL if idx % 2 == 0 else None
                ws_p.cell(row=r, column=1, value=ct).font = CELL_FONT
                ws_p.cell(row=r, column=1).border = THIN_BORDER
                if fill: ws_p.cell(row=r, column=1).fill = fill
                ws_p.cell(row=r, column=2, value=cd).font = CELL_FONT
                ws_p.cell(row=r, column=2).border = THIN_BORDER
                if fill: ws_p.cell(row=r, column=2).fill = fill
                ws_p.cell(row=r, column=3, value=iss['slicer_title']).font = CELL_FONT
                ws_p.cell(row=r, column=3).border = THIN_BORDER
                if fill: ws_p.cell(row=r, column=3).fill = fill
                ws_p.cell(row=r, column=4, value=iss['slicer_column']).font = Font(name='Meiryo', size=10, color='C00000')
                ws_p.cell(row=r, column=4).border = THIN_BORDER
                if fill: ws_p.cell(row=r, column=4).fill = fill
                r += 1

        ws_p.column_dimensions['A'].width = 40
        ws_p.column_dimensions['B'].width = 30
        ws_p.column_dimensions['C'].width = 25
        ws_p.column_dimensions['D'].width = 25

    xlsx_path = os.path.join(DIRECTORY, 'slicer_filter_audit.xlsx')
    wb.save(xlsx_path)
    print(f"Excel: {xlsx_path}")
    print(f"\nDone! {len(all_pages)} pages, {total_no} problem cards")


if __name__ == '__main__':
    main()
