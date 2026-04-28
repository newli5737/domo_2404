"""
Generate Vietnamese-English version of slicer filter audit report.
Reads the existing slicer_filter_audit.csv and re-generates with VN/EN labels.
"""
import csv, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

DIRECTORY = r"d:\DOMO\2404"

# Styles
HEADER_FONT = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
TITLE_FONT = Font(name='Segoe UI', size=14, bold=True, color='2F5496')
CELL_FONT = Font(name='Segoe UI', size=10)
CELL_SM = Font(name='Segoe UI', size=9, color='555555')
GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
RED_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
ALT_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
SLICER_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
BORDER = Border(
    left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7'))
CENTER = Alignment(horizontal='center', vertical='center')

# Read source CSV
src = os.path.join(DIRECTORY, 'slicer_filter_audit.csv')
rows = []
with open(src, 'r', encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        rows.append(r)

# Write VN/EN CSV
vn_csv = os.path.join(DIRECTORY, 'slicer_filter_audit_vn.csv')
vn_keys = ['page_id', 'page_name', 'slicer_name', 'filter_column', 'slicer_dataset',
            'card_id', 'card_name', 'card_dataset', 'responds', 'reason']
with open(vn_csv, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(vn_keys)
    for r in rows:
        reason_vn = {'同じDataset': 'Cùng dataset', 'カラム有り': 'Có column', 'カラム無し': 'Thiếu column'}
        w.writerow([
            r['page_id'], r['page_title'], r['slicer_title'], r['slicer_column'],
            r['slicer_dataset'], r['card_id'], r['card_title'], r['card_dataset'],
            r['responds'], reason_vn.get(r['reason'], r['reason'])
        ])
print(f"CSV (VN): {vn_csv} ({len(rows)} rows)")

# Group data
pages = {}
for r in rows:
    pid = r['page_id']
    if pid not in pages:
        pages[pid] = {'title': r['page_title'], 'rows': [], 'slicers': set()}
    pages[pid]['rows'].append(r)
    pages[pid]['slicers'].add(r['slicer_title'])

# Build Excel
wb = openpyxl.Workbook()

# ── Sheet 1: Tổng quan ──
ws = wb.active
ws.title = 'Tổng quan'
ws.sheet_properties.tabColor = '2F5496'

ws.merge_cells('A1:E1')
ws['A1'].value = 'Báo cáo kiểm tra Slicer Filter - Card Response Audit'
ws['A1'].font = TITLE_FONT
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:E2')
ws['A2'].value = 'Kiểm tra xem từng card trên page có phản ứng đúng với bộ lọc slicer hay không.'
ws['A2'].font = CELL_SM

row = 4
for ci, h in enumerate(['#', 'Tên Page', 'Page ID', 'Số Slicer', 'Card lỗi (❌)'], 1):
    c = ws.cell(row=row, column=ci, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER

total_err = 0
for i, (pid, pd) in enumerate(sorted(pages.items(), key=lambda x: int(x[0])), 1):
    row += 1
    n_err = sum(1 for r in pd['rows'] if r['responds'] == '❌')
    total_err += n_err
    ws.cell(row=row, column=1, value=i).font = CELL_FONT
    ws.cell(row=row, column=1).alignment = CENTER; ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2, value=pd['title']).font = CELL_FONT; ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=3, value=int(pid)).font = CELL_FONT
    ws.cell(row=row, column=3).alignment = CENTER; ws.cell(row=row, column=3).border = BORDER
    ws.cell(row=row, column=4, value=len(pd['slicers'])).font = CELL_FONT
    ws.cell(row=row, column=4).alignment = CENTER; ws.cell(row=row, column=4).border = BORDER
    c5 = ws.cell(row=row, column=5, value=n_err)
    c5.font = Font(name='Segoe UI', size=11, bold=True, color='C00000' if n_err > 0 else '107C10')
    c5.alignment = CENTER; c5.border = BORDER
    c5.fill = RED_FILL if n_err > 0 else GREEN_FILL

row += 1
ws.cell(row=row, column=2, value='Tổng cộng').font = Font(name='Segoe UI', size=11, bold=True)
ws.cell(row=row, column=2).border = BORDER
ws.cell(row=row, column=3).border = BORDER; ws.cell(row=row, column=4).border = BORDER
c5 = ws.cell(row=row, column=5, value=total_err)
c5.font = Font(name='Segoe UI', size=12, bold=True, color='C00000')
c5.alignment = CENTER; c5.border = BORDER

row += 2
ws.merge_cells(f'A{row}:E{row}')
ws[f'A{row}'].value = '✅ = Card phản ứng đúng với slicer  ／  ❌ = Card KHÔNG phản ứng (thiếu column trong dataset)'
ws[f'A{row}'].font = CELL_SM

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15

# ── Sheet 2: Card lỗi (chỉ ❌) ──
ws2 = wb.create_sheet('Card lỗi (không phản ứng)')
ws2.sheet_properties.tabColor = 'C00000'

no_rows = [r for r in rows if r['responds'] == '❌']
headers = ['#', 'Tên Page', 'Slicer', 'Column lọc', 'Dataset của Slicer',
           'Tên Card', 'Dataset của Card', 'Lý do']
for ci, h in enumerate(headers, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
ws2.row_dimensions[1].height = 28
ws2.freeze_panes = 'A2'

for i, r in enumerate(no_rows, 1):
    rr = i + 1
    fill = ALT_FILL if i % 2 == 0 else None
    vals = [i, r['page_title'], r['slicer_title'], r['slicer_column'], r['slicer_dataset'],
            r['card_title'], r['card_dataset'], 'Thiếu column trong dataset']
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(row=rr, column=ci, value=v)
        c.font = CELL_FONT; c.border = BORDER
        if fill: c.fill = fill
        if ci in (1,): c.alignment = CENTER

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 25
ws2.column_dimensions['D'].width = 25
ws2.column_dimensions['E'].width = 30
ws2.column_dimensions['F'].width = 45
ws2.column_dimensions['G'].width = 30
ws2.column_dimensions['H'].width = 25
if no_rows:
    ws2.auto_filter.ref = f'A1:H{len(no_rows)+1}'

# ── Sheet 3: Toàn bộ kết quả ──
ws3 = wb.create_sheet('Toàn bộ kết quả')
ws3.sheet_properties.tabColor = '548235'

headers = ['#', 'Tên Page', 'Slicer', 'Column lọc',
           'Tên Card', 'Dataset của Card', 'Phản ứng?', 'Lý do']
for ci, h in enumerate(headers, 1):
    c = ws3.cell(row=1, column=ci, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
ws3.row_dimensions[1].height = 28
ws3.freeze_panes = 'A2'

reason_vn = {'同じDataset': 'Cùng dataset', 'カラム有り': 'Có column', 'カラム無し': 'Thiếu column'}
for i, r in enumerate(rows, 1):
    rr = i + 1
    ok = r['responds'] == '✅'
    vals = [i, r['page_title'], r['slicer_title'], r['slicer_column'],
            r['card_title'], r['card_dataset'], r['responds'], reason_vn.get(r['reason'], r['reason'])]
    for ci, v in enumerate(vals, 1):
        c = ws3.cell(row=rr, column=ci, value=v)
        c.font = CELL_FONT; c.border = BORDER
        if ci in (1, 7): c.alignment = CENTER
        if ci == 7:
            c.fill = GREEN_FILL if ok else RED_FILL
            c.font = Font(name='Segoe UI', size=10, bold=True, color='107C10' if ok else 'C00000')

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 25
ws3.column_dimensions['E'].width = 45
ws3.column_dimensions['F'].width = 30
ws3.column_dimensions['G'].width = 10
ws3.column_dimensions['H'].width = 18
if rows:
    ws3.auto_filter.ref = f'A1:H{len(rows)+1}'

xlsx_path = os.path.join(DIRECTORY, 'slicer_audit_vn.xlsx')
wb.save(xlsx_path)
print(f"Excel (VN): {xlsx_path}")
print(f"Tổng: {len(rows)} checks, {len(no_rows)} ❌ lỗi, {len(rows)-len(no_rows)} ✅ OK")
